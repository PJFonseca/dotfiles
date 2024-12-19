import openpyxl
import xlwt
from datetime import datetime
import os
import glob

# Configuration
# Install: python split_xlsx_with_dates.py

OUTPUT_DIR = "output_chunks"
HEADER_ROWS = 4  # Number of rows to keep as the header
CHUNK_SIZE = 1000  # Number of rows per chunk (excluding the header)

# Search for the first .xlsx file in the current directory
xlsx_files = glob.glob("*.xlsx")
if not xlsx_files:
    print("No .xlsx files found in the current directory.")
    exit(1)

INPUT_FILE = xlsx_files[0]  # Use the first .xlsx file found
print(f"Found input file: {INPUT_FILE}")

# Ensure the output directory exists
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Load the workbook
print(f"Loading {INPUT_FILE}...")
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
sheet = wb.active

# Extract the header rows
print("Extracting header rows...")
header = []
for row in sheet.iter_rows(min_row=1, max_row=HEADER_ROWS, values_only=False):
    header.append([cell for cell in row])

# Function to convert dates in DD/MM/YYYY to YYYY-MM-DD
def convert_date(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return value

# Process the remaining rows
print("Splitting data into chunks...")
row_count = sheet.max_row
current_chunk = 1
for start_row in range(HEADER_ROWS + 1, row_count + 1, CHUNK_SIZE):
    # Create a new workbook for the chunk (xlwt, for .xls)
    new_wb = xlwt.Workbook()
    new_ws = new_wb.add_sheet("Sheet1")

    # Copy the header rows to the new .xls file
    for r, header_row in enumerate(header):
        for c, cell in enumerate(header_row):
            new_ws.write(r, c, convert_date(cell.value) if cell.value else None)

    # Copy rows for this chunk
    end_row = min(start_row + CHUNK_SIZE - 1, row_count)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=False)):
        for col_idx, cell in enumerate(row):
            new_ws.write(row_idx + HEADER_ROWS, col_idx, convert_date(cell.value) if cell.value else None)

    # Save the chunk to a new file
    chunk_filename = os.path.join(OUTPUT_DIR, f"chunk_{current_chunk:03d}.xls")
    new_wb.save(chunk_filename)
    print(f"Saved {chunk_filename}")
    current_chunk += 1

print("Done! Split files are saved in the 'output_chunks' directory.")
